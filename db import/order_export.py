import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
from openpyxl.utils import get_column_letter

# Конфигурация базы данных
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '300132',
    'db': 'flask_auth',
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# Производители для группировки
MANUFACTURERS = ['Свежая выпечка', 'Хлебный двор']

# Директория для сохранения файлов
OUTPUT_DIR = 'exports'
os.makedirs(OUTPUT_DIR, exist_ok=True)


def get_orders_by_manufacturer(manufacturer):
    """Получает заказы для конкретного производителя, сгруппированные по магазинам"""
    connection = pymysql.connect(**DB_CONFIG)
    try:
        with connection.cursor() as cursor:
            sql = """
            SELECT 
                s.name AS shop_name,
                s.address AS shop_address,
                GROUP_CONCAT(CONCAT(p.name, ' (', oi.quantity, ' шт.)') SEPARATOR '\n') AS products_list,
                SUM(oi.quantity * p.price) AS total_amount
            FROM orders o
            JOIN order_items oi ON oi.order_id = o.id
            JOIN products p ON p.id = oi.product_id
            JOIN shops s ON s.id = o.shop_id
            WHERE p.manufacturer = %s
            GROUP BY s.id
            ORDER BY s.name
            """
            cursor.execute(sql, (manufacturer,))
            return cursor.fetchall()
    finally:
        connection.close()


def create_excel_for_manufacturer(manufacturer, orders):
    """Создает Excel файл для производителя с группировкой по магазинам"""
    if not orders:
        print(f"Нет заказов для производителя: {manufacturer}")
        return

    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Заголовок файла
    ws.cell(row=1, column=1, value=f"Заказы производителя: {manufacturer}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    # Заголовки таблицы
    headers = ["Магазин", "Адрес", "Товары (количество)", "Общая сумма"]

    # Стили
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(vertical='top', wrap_text=True)

    # Записываем заголовки таблицы
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Заполняем данными
    row_num = 3
    for order in orders:
        ws.cell(row=row_num, column=1, value=order['shop_name']).border = border
        ws.cell(row=row_num, column=2, value=order['shop_address']).border = border
        ws.cell(row=row_num, column=3, value=order['products_list']).border = border
        ws.cell(row=row_num, column=3).alignment = wrap_alignment
        ws.cell(row=row_num, column=4, value=order['total_amount']).border = border
        row_num += 1

    # Автонастройка ширины столбцов
    for col_idx, col_name in enumerate(headers, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        # Проверяем все ячейки в столбце
        for row in ws.iter_rows(min_row=2, max_row=row_num - 1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                try:
                    # Для объединенных ячеек используем значение только из первой ячейки
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass

        # Разные настройки ширины для разных столбцов
        if col_name == "Товары (количество)":
            adjusted_width = min(max_length + 2, 50) * 1.2
        else:
            adjusted_width = (max_length + 2) * 1.2

        ws.column_dimensions[column_letter].width = adjusted_width

    # Устанавливаем высоту строк для столбца с товарами
    for row_idx in range(3, row_num):
        products = ws.cell(row=row_idx, column=3).value
        if products:
            # Высота строки зависит от количества строк в товарах
            line_count = products.count('\n') + 1
            ws.row_dimensions[row_idx].height = min(15 * line_count, 150)  # Ограничиваем максимальную высоту

    # Форматирование чисел
    for row in ws.iter_rows(min_row=3, max_col=4, max_row=row_num - 1):
        # Общая сумма
        if isinstance(row[3].value, (int, float)):
            row[3].number_format = '#,##0'

    # Итоги
    total_row = row_num
    ws.cell(row=total_row, column=3, value="Итого:").font = Font(bold=True)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=4,
            value=f"=SUM(D3:D{row_num - 1})").font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '#,##0'

    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{manufacturer}_заказы_{timestamp}.xlsx".replace(" ", "_")
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    print(f"Создан файл: {filepath}")
    print(f"Количество магазинов: {len(orders)}")
    print(f"Общая сумма: {sum(order['total_amount'] for order in orders):,.0f} тенге\n")


def main():
    print("Начало экспорта заказов...\n")

    for manufacturer in MANUFACTURERS:
        print(f"Обработка производителя: {manufacturer}")
        orders = get_orders_by_manufacturer(manufacturer)
        create_excel_for_manufacturer(manufacturer, orders)

    print("Экспорт завершен!")


if __name__ == "__main__":
    main()